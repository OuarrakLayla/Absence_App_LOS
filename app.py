from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
import sqlite3, io
from datetime import datetime
from functools import wraps
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'gestiabsence-secret-2026'
DB_PATH = 'absences.db'

ADMIN = {
    'username': 'admin',
    'password': 'Admin@2026'
}

# ── Auth ───────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def deco(*a, **kw):
        if not session.get('logged_in'):
            return redirect(url_for('login_page'))
        return f(*a, **kw)
    return deco

def api_auth(f):
    @wraps(f)
    def deco(*a, **kw):
        if not session.get('logged_in'):
            return jsonify({'error': 'Non autorisé'}), 401
        return f(*a, **kw)
    return deco

# ── DB ─────────────────────────────────────────────────────
def get_db():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    return c

def init_db():
    c = get_db()
    c.execute('''CREATE TABLE IF NOT EXISTS absences (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        nom TEXT NOT NULL,
        prenom TEXT NOT NULL,
        matricule TEXT NOT NULL,
        classe TEXT NOT NULL,
        email TEXT,
        date TEXT NOT NULL,
        type_absence TEXT NOT NULL,
        motif TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')
    if c.execute('SELECT COUNT(*) FROM absences').fetchone()[0] == 0:
        rows = [
            ('Benali','Youssef','ETU001','INFO-3A','youssef.benali@uni.ma','2026-03-01','Non justifiée','Absent sans motif'),
            ('Alaoui','Fatima','ETU002','INFO-3A','fatima.alaoui@uni.ma','2026-03-03','Justifiée','Maladie'),
            ('Tazi','Omar','ETU003','MATH-2B','omar.tazi@uni.ma','2026-03-05','Non justifiée',''),
            ('Cherif','Nadia','ETU004','MATH-2B','nadia.cherif@uni.ma','2026-03-07','Justifiée','Convocation administrative'),
            ('Hajji','Khalid','ETU005','PHYS-1C','khalid.hajji@uni.ma','2026-03-08','Non justifiée','Retard non signalé'),
            ('Idrissi','Samira','ETU006','INFO-3A','samira.idrissi@uni.ma','2026-03-10','Justifiée','Urgence familiale'),
            ('Mrani','Hassan','ETU007','PHYS-1C','hassan.mrani@uni.ma','2026-03-11','Non justifiée',''),
            ('Bouazza','Leila','ETU008','MATH-2B','leila.bouazza@uni.ma','2026-03-12','Justifiée','Maladie'),
            ('Naciri','Rachid','ETU009','INFO-3A','rachid.naciri@uni.ma','2026-02-15','Non justifiée',''),
            ('Fassi','Amina','ETU010','PHYS-1C','amina.fassi@uni.ma','2026-02-20','Justifiée','Maladie chronique'),
            ('Bakkali','Mehdi','ETU011','INFO-3A','mehdi.bakkali@uni.ma','2026-01-10','Non justifiée',''),
            ('Ziani','Sofia','ETU012','MATH-2B','sofia.ziani@uni.ma','2026-01-18','Justifiée','Accident'),
        ]
        c.executemany(
            'INSERT INTO absences (nom,prenom,matricule,classe,email,date,type_absence,motif) VALUES (?,?,?,?,?,?,?,?)',
            rows
        )
    c.commit()
    c.close()

# ── Auth routes ────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if session.get('logged_in'):
        return redirect('/')
    if request.method == 'POST':
        d = request.json or request.form
        if d.get('username') == ADMIN['username'] and d.get('password') == ADMIN['password']:
            session['logged_in'] = True
            session['username']  = ADMIN['username']
            if request.is_json:
                return jsonify({'success': True})
            return redirect('/')
        if request.is_json:
            return jsonify({'success': False, 'error': 'Identifiants incorrects'}), 401
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect('/login')

# ── Pages ──────────────────────────────────────────────────
@app.route('/')
@login_required
def dashboard():
    return render_template('dashboard.html', active='dashboard', now=datetime.now().strftime('%d %B %Y'))

@app.route('/ajouter')
@login_required
def ajouter():
    return render_template('ajouter.html', active='ajouter', now=datetime.now().strftime('%d %B %Y'))

@app.route('/statistiques')
@login_required
def statistiques():
    return render_template('statistiques.html', active='stats', now=datetime.now().strftime('%d %B %Y'))

# ── API ────────────────────────────────────────────────────
@app.route('/api/absences', methods=['GET'])
@api_auth
def get_absences():
    c = get_db()
    q = 'SELECT * FROM absences WHERE 1=1'
    p = []
    for k, col in [('type', 'type_absence'), ('classe', 'classe')]:
        v = request.args.get(k)
        if v:
            q += f' AND {col}=?'
            p.append(v)
    e = request.args.get('etudiant')
    if e:
        q += ' AND (nom LIKE ? OR prenom LIKE ? OR matricule LIKE ?)'
        p += [f'%{e}%'] * 3
    d1, d2 = request.args.get('date_debut'), request.args.get('date_fin')
    if d1:
        q += ' AND date>=?'; p.append(d1)
    if d2:
        q += ' AND date<=?'; p.append(d2)
    rows = c.execute(q + ' ORDER BY date DESC', p).fetchall()
    c.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/absences', methods=['POST'])
@api_auth
def add_absence():
    d = request.json
    c = get_db()
    c.execute(
        'INSERT INTO absences (nom,prenom,matricule,classe,email,date,type_absence,motif) VALUES (?,?,?,?,?,?,?,?)',
        (d['nom'], d['prenom'], d['matricule'], d['classe'],
         d.get('email', ''), d['date'], d['type_absence'], d.get('motif', ''))
    )
    c.commit()
    c.close()
    return jsonify({'success': True})

@app.route('/api/absences/<int:id>', methods=['DELETE'])
@api_auth
def delete_absence(id):
    c = get_db()
    c.execute('DELETE FROM absences WHERE id=?', (id,))
    c.commit()
    c.close()
    return jsonify({'success': True})

@app.route('/api/stats')
@api_auth
def get_stats():
    c = get_db()
    total = c.execute('SELECT COUNT(*) FROM absences').fetchone()[0]
    just  = c.execute("SELECT COUNT(*) FROM absences WHERE type_absence='Justifiée'").fetchone()[0]
    bc = c.execute("""SELECT classe, COUNT(*) as total,
        SUM(CASE WHEN type_absence='Justifiée' THEN 1 ELSE 0 END) as justifiees,
        SUM(CASE WHEN type_absence='Non justifiée' THEN 1 ELSE 0 END) as non_justifiees
        FROM absences GROUP BY classe ORDER BY total DESC""").fetchall()
    bs = c.execute("""SELECT nom||' '||prenom as etudiant, matricule, COUNT(*) as total
        FROM absences GROUP BY matricule ORDER BY total DESC LIMIT 10""").fetchall()
    bt = c.execute('SELECT type_absence, COUNT(*) as count FROM absences GROUP BY type_absence').fetchall()
    bm = c.execute("""SELECT substr(date,1,7) as mois, COUNT(*) as total,
        SUM(CASE WHEN type_absence='Justifiée' THEN 1 ELSE 0 END) as justifiees,
        SUM(CASE WHEN type_absence='Non justifiée' THEN 1 ELSE 0 END) as non_justifiees
        FROM absences GROUP BY mois ORDER BY mois""").fetchall()
    cls = c.execute('SELECT DISTINCT classe FROM absences ORDER BY classe').fetchall()
    c.close()
    return jsonify({
        'total': total, 'justifiees': just, 'non_justifiees': total - just,
        'by_class':   [dict(r) for r in bc],
        'by_student': [dict(r) for r in bs],
        'by_type':    [dict(r) for r in bt],
        'by_month':   [dict(r) for r in bm],
        'classes':    [r['classe'] for r in cls]
    })

@app.route('/api/export/excel')
@api_auth
def export_excel():
    c = get_db()
    rows = c.execute('SELECT * FROM absences ORDER BY date DESC').fetchall()
    c.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Absences"
    thin = Side(style='thin', color="E2E8F0")
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells('A1:H1')
    ws['A1'] = 'GESTIABSENCE — Rapport des Absences'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill('solid', fgColor='0F172A')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 34

    ws.merge_cells('A2:H2')
    ws['A2'] = f"Généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}  •  {len(rows)} absences"
    ws['A2'].font = Font(name='Arial', size=10, color='94A3B8', italic=True)
    ws['A2'].fill = PatternFill('solid', fgColor='1E293B')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    for col, h in enumerate(['N°','Nom','Prénom','Matricule','Classe','Date',"Type d'absence",'Motif'], 1):
        cl = ws.cell(row=4, column=col, value=h)
        cl.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
        cl.fill = PatternFill('solid', fgColor='6366F1')
        cl.alignment = Alignment(horizontal='center', vertical='center')
        cl.border = brd
    ws.row_dimensions[4].height = 22

    for i, r in enumerate(rows):
        rn  = 5 + i
        bg  = 'FFFFFF' if i % 2 == 0 else 'F8FAFC'
        is_j = r['type_absence'] == 'Justifiée'
        for col, val in enumerate([i+1, r['nom'], r['prenom'], r['matricule'], r['classe'], r['date'], r['type_absence'], r['motif'] or ''], 1):
            cl = ws.cell(row=rn, column=col, value=val)
            cl.font = Font(name='Arial', size=10)
            cl.fill = PatternFill('solid', fgColor=bg)
            cl.alignment = Alignment(horizontal='center' if col in [1,4,5,6] else 'left', vertical='center')
            cl.border = brd
            if col == 7:
                cl.font = Font(name='Arial', size=10, bold=True, color='065F46' if is_j else '991B1B')
                cl.fill = PatternFill('solid', fgColor='D1FAE5' if is_j else 'FEE2E2')
        ws.row_dimensions[rn].height = 18

    for i, w in enumerate([5, 15, 15, 12, 12, 14, 18, 28], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'A5'

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'absences_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)